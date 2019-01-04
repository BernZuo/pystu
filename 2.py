from openpyxl import load_workbook
import pymysql
import arrow
def doExcel(file_name,sheet_name):
    wd =load_workbook(file_name)
    sheet = wd[sheet_name]
    data = []
    for i in range(2, sheet.max_row + 1):
        param = {}
        for j in range(1, sheet.max_column + 1):
            # 获取第一行，多少列的值
            key = sheet.cell(1, j).value
            key=key.strip()
            # 获取多少行，多少列的值
            value = sheet.cell(i, j).value
            if isinstance(value,str):
                value = value.strip()
            param[key] = value
        data.append(param)
    return data

def mySQL_01():
    config={'host':'192.168.1.50',
    'user':"root",
    'password':"root",
    'port':3306}
    #连接打开数据库
    conn=pymysql.connect(**config)

    # 使用cursor()方法获取操作游标
    cur = conn.cursor()

    try:
        # 创建数据库
        createSql = ''
        DB_NAME = 'bill'
        TB_NAME = 'dd' + arrow.now().shift(months=-1).format("YYYYMM") + 'bill'
        cur.execute('CREATE DATABASE IF NOT EXISTS %s' %DB_NAME)
        conn.select_db(DB_NAME)
        #调用自定义doExcel方法，获取文件表头
        a = doExcel('大德12月服务接口调用量.xlsx', 'Sheet1')
        for key in a[0].keys():
            if isinstance(a[0][key],str):
                createSql = createSql + key + ' varchar(255)' + ','
            else:
                createSql = createSql + key + ' int(11)' + ','
        createSql = 'CREATE TABLE IF NOT EXISTS %s(%s)DEFAULT CHARSET=utf8' %(TB_NAME,createSql[:-1])
        cur.execute(createSql)
        for i in a:
            data = i
            #print(i)
            sql = "INSERT INTO {}(客户标识,客户名称,数据源,运营商,服务项编码,接口,接口名称,承载平台,产品名称,总和调用量) VALUES('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}')" .format(
                TB_NAME,data['客户标识'], data['客户名称'], data['数据源'], data['运营商'], data['服务项编码'], data['接口'], data['接口名称'], data['承载平台'],data['产品名称'], data['总和调用量'])
            print(sql)
            cur.execute(sql)
        #获取结果  1）获取单条  2）获取多条
        # res=cur.fetchone()
        # res=cur.fetchall()
        conn.commit()

    except:
        import traceback
        traceback.print_exc()
        conn.rollback()
    finally:
    #关掉游标，关掉连接
        cur.close()
        conn.close()

if __name__ == '__main__':
        #mySql_conn()
        mySQL_01()
